#!/usr/bin/env python3
"""
Notes Vault v2.0 — A fast, clean CLI notes app
================================================
Features:
  • Create notes with title + body (multi-line editor)
  • Note templates (Meeting, Journal, To-Do, Blank)
  • Browse notes by date, category, or view all (with sort options)
  • Full-text search across every note ever written
  • Pin important notes so they're always visible
  • Archive & Trash with recovery
  • Duplicate notes
  • Edit or append to existing notes
  • Categories with color coding
  • Export to Markdown, plain text, CSV, and Excel (.xlsx)
  • Configurable save location
  • Word/character counts and usage stats
  • Undo last 10 actions

Config: ~/.notes_vault_config.json  (stores save path + export dir)
Data:   ~/.notes_vault.json         (or custom path via settings)
"""

from __future__ import annotations

import copy
import csv
import json
import os
import re
import shutil
import shlex
import subprocess
import sys
import tempfile
import textwrap
import time as time_module
from datetime import datetime, date, timedelta
from typing import List, Optional, Dict, Any, Tuple

# ── Config (separate from data so moving data file doesn't orphan config) ──

CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".notes_vault_config.json")

DEFAULT_CONFIG = {
    "data_path": os.path.join(os.path.expanduser("~"), ".notes_vault.json"),
    "export_dir": os.path.join(os.path.expanduser("~"), "notes_exports"),
}

def load_config() -> Dict[str, str]:
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            for k, v in DEFAULT_CONFIG.items():
                if k not in cfg:
                    cfg[k] = v
            return cfg
        except Exception:
            pass
    return copy.deepcopy(DEFAULT_CONFIG)

def save_config(cfg: Dict[str, str]) -> None:
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2)

CONFIG = load_config()

def data_path() -> str:
    return CONFIG["data_path"]

def export_dir() -> str:
    return CONFIG["export_dir"]


# ════════════════════════════════════════════════════════════════
#  TERMINAL UTILITIES
# ════════════════════════════════════════════════════════════════

def term_width() -> int:
    return shutil.get_terminal_size((80, 24)).columns

def is_tty() -> bool:
    return sys.stdout.isatty()

def c(text: str, code: str) -> str:
    return f"\033[{code}m{text}\033[0m" if is_tty() else text

def clear():
    os.system("cls" if os.name == "nt" else "clear")

def pause(msg: str = "\n  Press Enter to continue..."):
    input(msg)

def wrap_text(text: str, width: int = 72, indent: str = "    ") -> str:
    lines = text.split("\n")
    wrapped = []
    for line in lines:
        if line.strip() == "":
            wrapped.append("")
        else:
            wrapped.extend(textwrap.wrap(line, width=width, initial_indent=indent,
                                         subsequent_indent=indent))
    return "\n".join(wrapped)


# ── Visual Components ──────────────────────────────────────────

def _bw() -> int:
    return min(term_width(), 74)

def draw_header(title: str, subtitle: str = "") -> None:
    w = _bw()
    inner = w - 2
    print()
    print(c(f"  ╔{'═' * inner}╗", "90"))
    pad_total = inner - len(title) - 2
    pad_l = pad_total // 2
    pad_r = pad_total - pad_l
    print(c("  ║", "90") + " " * pad_l + f" {c(title, '1;37')} " + " " * pad_r + c("║", "90"))
    if subtitle:
        s_pad = inner - len(subtitle) - 2
        sl = s_pad // 2
        sr = s_pad - sl
        print(c("  ║", "90") + " " * sl + f" {c(subtitle, '90')} " + " " * sr + c("║", "90"))
    print(c(f"  ╚{'═' * inner}╝", "90"))
    print()

def draw_section(title: str) -> None:
    w = _bw() - 4
    pad = w - len(title) - 3
    print(f"  {c('─── ', '90')}{c(title, '1;37')}{c(' ' + '─' * max(pad, 1), '90')}")
    print()

def draw_divider() -> None:
    print(c(f"  {'─' * (_bw() - 4)}", "90"))

def draw_menu(options: List[Tuple[str, str]], columns: int = 2) -> None:
    print()
    draw_divider()
    print()
    if columns == 1:
        for num, label in options:
            print(f"  {c(f'  [{num}]', '36')}  {label}")
    else:
        col_w = (_bw() - 4) // 2
        rows = (len(options) + 1) // 2
        for i in range(rows):
            li, ri = i, i + rows
            left = right = ""
            if li < len(options):
                n, l = options[li]
                left = f"  {c(f'[{n}]', '36')}  {l}"
            if ri < len(options):
                n, l = options[ri]
                right = f"{c(f'[{n}]', '36')}  {l}"
            vis = re.sub(r'\033\[[^m]*m', '', left)
            pad = col_w - len(vis)
            print(f"{left}{' ' * max(pad, 2)}{right}")
    print()

def draw_prompt() -> str:
    return input(f"  {c('›', '36')} ").strip()

def draw_inline_menu(options: List[Tuple[str, str]]) -> None:
    print()
    parts = [f"{c(f'[{n}]', '36')} {l}" for n, l in options]
    print(f"  {'    '.join(parts)}")
    print()


# ════════════════════════════════════════════════════════════════
#  DATA LAYER
# ════════════════════════════════════════════════════════════════

DEFAULT_DATA: Dict[str, Any] = {
    "notes": [],
    "archive": [],
    "trash": [],
    "categories": ["General", "Work", "Personal", "Ideas"],
    "templates": [
        {"name": "Meeting Notes", "body": "Attendees:\n\nAgenda:\n\nDiscussion:\n\nAction Items:\n"},
        {"name": "Journal Entry", "body": "How I'm feeling:\n\nWhat happened today:\n\nWhat I'm grateful for:\n"},
        {"name": "To-Do List", "body": "[ ] \n[ ] \n[ ] \n[ ] \n[ ] \n"},
    ],
    "settings": {
        "default_category": "General",
        "editor_hint": True,
        "trash_days": 30,
        "use_external_editor": False,
    },
    "undo_stack": [],
}

def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")

def today_str() -> str:
    return date.today().isoformat()

def load_data() -> Dict[str, Any]:
    p = data_path()
    if not os.path.exists(p):
        return copy.deepcopy(DEFAULT_DATA)
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
        for k, v in DEFAULT_DATA.items():
            if k not in data:
                data[k] = copy.deepcopy(v)
        for k, v in DEFAULT_DATA["settings"].items():
            if k not in data.get("settings", {}):
                data["settings"][k] = v
        return data
    except Exception:
        return copy.deepcopy(DEFAULT_DATA)

def save_data(data: Dict[str, Any]) -> None:
    p = data_path()
    os.makedirs(os.path.dirname(p) or ".", exist_ok=True)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def next_id(data: Dict[str, Any]) -> int:
    all_ids = [n.get("id", 0) for n in data.get("notes", [])]
    all_ids += [n.get("id", 0) for n in data.get("archive", [])]
    all_ids += [n.get("id", 0) for n in data.get("trash", [])]
    return (max(all_ids) + 1) if all_ids else 1

def auto_purge_trash(data: Dict[str, Any]) -> None:
    """Remove trash items older than trash_days."""
    days = data.get("settings", {}).get("trash_days", 30)
    cutoff = (datetime.now() - timedelta(days=days)).isoformat()
    before = len(data.get("trash", []))
    data["trash"] = [n for n in data.get("trash", []) if (n.get("trashed_at") or "") > cutoff]
    purged = before - len(data["trash"])
    if purged > 0:
        save_data(data)


# ── Undo ───────────────────────────────────────────────────────

MAX_UNDO = 10

def push_undo(data: Dict[str, Any], desc: str) -> None:
    snapshot = {
        "desc": desc, "ts": now_iso(),
        "notes": copy.deepcopy(data.get("notes", [])),
        "archive": copy.deepcopy(data.get("archive", [])),
        "trash": copy.deepcopy(data.get("trash", [])),
    }
    stack = data.setdefault("undo_stack", [])
    stack.append(snapshot)
    if len(stack) > MAX_UNDO:
        data["undo_stack"] = stack[-MAX_UNDO:]

def do_undo(data: Dict[str, Any]) -> Optional[str]:
    stack = data.get("undo_stack", [])
    if not stack:
        return None
    snap = stack.pop()
    data["notes"] = snap["notes"]
    data["archive"] = snap["archive"]
    data["trash"] = snap.get("trash", [])
    save_data(data)
    return snap["desc"]


# ════════════════════════════════════════════════════════════════
#  NOTE FORMATTING
# ════════════════════════════════════════════════════════════════

CATEGORY_COLORS = {
    "general": "37", "work": "33", "personal": "35", "ideas": "36",
    "journal": "34", "meetings": "31", "goals": "32",
}

def cat_color(category: str) -> str:
    return CATEGORY_COLORS.get(category.lower(), "37")


def parse_tags(raw: str) -> List[str]:
    tags = [t.strip() for t in raw.split(",") if t.strip()]
    seen = set()
    unique = []
    for tag in tags:
        key = tag.lower()
        if key not in seen:
            seen.add(key)
            unique.append(tag)
    return unique


def format_tags(tags: List[str]) -> str:
    if not tags:
        return ""
    return ", ".join(tags)


def extract_links(body: str) -> Tuple[List[int], List[str]]:
    ids = [int(n) for n in re.findall(r"\B#(\d+)\b", body)]
    titles = [t.strip() for t in re.findall(r"\[\[([^\]]+)\]\]", body)]
    return ids, titles


def resolve_link_targets(data: Dict[str, Any], body: str) -> List[Dict[str, Any]]:
    ids, titles = extract_links(body)
    targets = []
    all_notes = data.get("notes", []) + data.get("archive", []) + data.get("trash", [])
    for nid in ids:
        note = next((n for n in all_notes if n.get("id") == nid), None)
        if note:
            targets.append({"label": f"#{nid} {note.get('title', 'Untitled')}", "note": note})
    for title in titles:
        note = next((n for n in all_notes if (n.get("title") or "").lower() == title.lower()), None)
        if note:
            targets.append({"label": f"[[{title}]] → #{note.get('id')}", "note": note})
    return targets


def open_note_target(data: Dict[str, Any], note: Dict[str, Any]) -> None:
    nid = note.get("id")
    if any(n.get("id") == nid for n in data.get("notes", [])):
        view_note(data, note)
    elif any(n.get("id") == nid for n in data.get("archive", [])):
        view_archived_note(data, note)
    elif any(n.get("id") == nid for n in data.get("trash", [])):
        view_trashed_note(data, note)

def format_note_line(note: Dict[str, Any], show_preview: bool = True) -> str:
    nid = note.get("id", 0)
    title = note.get("title", "Untitled")
    cat = note.get("category", "General")
    pinned = note.get("pinned", False)
    created = note.get("created_at", "")[:10]
    updated = note.get("updated_at")
    body = note.get("body", "")
    tags = note.get("tags", [])
    wc = len(body.split()) if body else 0

    pin = c("📌", "1;33") if pinned else "  "
    edited = c(" ✎", "90") if (updated and updated != note.get("created_at")) else ""
    tag_display = f" {c('· ' + ', '.join(tags), '90')}" if tags else ""
    line = f"    {pin} {c(f'#{nid:<4}', '1;37')}  {c(created, '90')}  {c(f'[{cat}]', cat_color(cat)):<22} {title}{edited}{tag_display}  {c(f'{wc}w', '90')}"

    if show_preview and body:
        preview = body.replace("\n", " ").strip()
        if len(preview) > 55:
            preview = preview[:52] + "..."
        line += f"\n         {c(preview, '90')}"
    return line


def display_note_full(note: Dict[str, Any]) -> None:
    w = _bw()
    inner = w - 4
    title = note.get("title", "Untitled")
    body = note.get("body", "")
    cat = note.get("category", "General")
    created = note.get("created_at", "")
    updated = note.get("updated_at")
    pinned = note.get("pinned", False)
    nid = note.get("id", 0)
    tags = note.get("tags", [])
    wc = len(body.split()) if body else 0
    cc = len(body)
    pin_mark = " 📌" if pinned else ""

    print()
    print(c(f"  ┌{'─' * inner}┐", "90"))
    print(f"  {c('│', '90')} {c(f'#{nid}', '36')} {c(title, '1;37')}{pin_mark}")
    meta = f"{c(f'[{cat}]', cat_color(cat))}  {c(created[:16], '90')}"
    if updated and updated != created:
        meta += c(f"  ·  edited {updated[:16]}", "90")
    print(f"  {c('│', '90')} {meta}")
    print(c(f"  └{'─' * inner}┘", "90"))

    tags_line = format_tags(tags) or "None"
    updated_line = updated[:16] if updated else "—"
    print(f"  {c('Category:', '90')} {cat}")
    print(f"  {c('Tags:', '90')} {tags_line}")
    print(f"  {c('Created:', '90')} {created[:16]}")
    print(f"  {c('Updated:', '90')} {updated_line}")

    if body:
        print()
        print(wrap_text(body, width=inner - 2))
    else:
        print(c("\n    (empty note)", "90"))

    print()
    footer = f"{wc} words · {cc} chars"
    fw = _bw() - 4
    pad = fw - len(footer) - 2
    l = pad // 2
    r = pad - l
    print(c(f"  {'─' * l} {footer} {'─' * r}", "90"))


# ════════════════════════════════════════════════════════════════
#  MULTI-LINE EDITOR
# ════════════════════════════════════════════════════════════════

def multiline_input(existing: str = "", hint: bool = True) -> str:
    if hint:
        print(c("    Type your note. Commands on their own line:", "90"))
        print(c("      :done (or :d) = save    :cancel = discard    :clear = start over", "90"))
        print()
    lines = existing.split("\n") if existing else []
    if existing:
        print(c("    (Existing content loaded. New text appends below.)", "90"))
        for i, line in enumerate(lines):
            print(c(f"    {i+1:>3} │ ", "90") + line)
        print()
    new_lines = []
    line_num = len(lines) + 1
    while True:
        try:
            line = input(c(f"    {line_num:>3} │ ", "90"))
        except (EOFError, KeyboardInterrupt):
            print()
            break
        cmd = line.strip().lower()
        if cmd in (":done", ":d"):
            break
        if cmd == ":cancel":
            return "__CANCEL__"
        if cmd == ":clear":
            new_lines, lines, line_num = [], [], 1
            print(c("    (cleared)", "90"))
            continue
        new_lines.append(line)
        line_num += 1
    if existing and new_lines:
        return existing.rstrip("\n") + "\n" + "\n".join(new_lines)
    elif new_lines:
        return "\n".join(new_lines)
    return existing


def build_meeting_note_body() -> str:
    print(c("    Fill in meeting sections. Use :done on a blank line to finish each section.", "90"))
    print()
    attendees = input("    Attendees (comma-separated): ").strip()
    print()
    print(c("    Agenda:", "90"))
    agenda = multiline_input(hint=False)
    print()
    print(c("    Discussion:", "90"))
    discussion = multiline_input(hint=False)
    print()
    print(c("    Action Items:", "90"))
    action_items = multiline_input(hint=False)
    return (
        "Attendees:\n"
        f"{attendees}\n\n"
        "Agenda:\n"
        f"{agenda}\n\n"
        "Discussion:\n"
        f"{discussion}\n\n"
        "Action Items:\n"
        f"{action_items}\n"
    )


# ════════════════════════════════════════════════════════════════
#  NOTE OPERATIONS
# ════════════════════════════════════════════════════════════════

def create_note(data: Dict[str, Any]) -> None:
    clear()
    draw_header("✏️  New Note", "Create a new note with a title and body")

    # Template selection
    templates = data.get("templates", [])
    template_name = ""
    template_body = ""
    if templates:
        draw_section("Start from template?")
        print(f"    {c('[0]', '36')}  Blank note")
        for i, t in enumerate(templates, 1):
            print(f"    {c(f'[{i}]', '36')}  {t['name']}")
        print()
        tmpl_choice = draw_prompt()
        if tmpl_choice.isdigit() and 1 <= int(tmpl_choice) <= len(templates):
            selected_template = templates[int(tmpl_choice) - 1]
            template_name = selected_template.get("name", "")
            template_body = selected_template.get("body", "")
        print()

    title = input("    Title: ").strip()
    if not title:
        print("    No title — cancelled.")
        pause()
        return

    cats = data.get("categories", [])
    default_cat = data.get("settings", {}).get("default_category", "General")
    if cats:
        print(f"\n    Categories: {'  '.join(c(f'[{ct}]', cat_color(ct)) for ct in cats)}")
    category = input(f"    Category [{default_cat}]: ").strip() or default_cat
    if category not in cats:
        cats.append(category)
        data["categories"] = cats

    raw_tags = input("    Tags (comma-separated, optional): ").strip()
    tags = parse_tags(raw_tags) if raw_tags else []

    print()
    if category not in cats:
        cats.append(category)
        data["categories"] = cats

    print()
    if template_name.lower() == "meeting notes":
        body = build_meeting_note_body()
    else:
        initial = template_body if templates and template_body else ""
        if initial:
            print(c(f"    Template loaded: {template_name}\n", "36"))
        body = body_input(existing=initial, hint=data.get("settings", {}).get("editor_hint", True), settings=data.get("settings", {}))
        body = multiline_input(existing=initial, hint=data.get("settings", {}).get("editor_hint", True))
    if body == "__CANCEL__":
        print("    Cancelled.")
        pause()
        return

    note = {
        "id": next_id(data), "title": title, "body": body,
        "category": category, "tags": tags, "created_at": now_iso(),
        "updated_at": now_iso(), "pinned": False,
    }
    push_undo(data, f"Create note #{note['id']}")
    data.setdefault("notes", []).append(note)
    save_data(data)
    print(f"\n    {c('✓ Saved', '1;32')} — #{note['id']} \"{title}\"")
    pause()


def view_note(data: Dict[str, Any], note: Dict[str, Any]) -> None:
    while True:
        clear()
        draw_header(f"📄 Note #{note.get('id', 0)}")
        display_note_full(note)
        link_targets = resolve_link_targets(data, note.get("body", ""))
        options = [
            ("1", "Edit"), ("2", "Append"), ("3", "Pin/Unpin"),
            ("4", "Duplicate"), ("5", "Archive"), ("6", "Delete"), ("0", "Back"),
        ]
        if link_targets:
            options.insert(6, ("7", "Open link"))
        draw_inline_menu(options)
        ch = draw_prompt()
        if ch == "1":   edit_note(data, note)
        elif ch == "2": append_to_note(data, note)
        elif ch == "3": toggle_pin(data, note)
        elif ch == "4": duplicate_note(data, note); return
        elif ch == "5": archive_note(data, note); return
        elif ch == "6":
            if trash_note(data, note):
                return
        elif ch == "7" and link_targets:
            print()
            for idx, target in enumerate(link_targets, 1):
                print(f"    {c(f'[{idx}]', '36')} {target['label']}")
            raw = input("    Open link #: ").strip()
            if raw.isdigit():
                sel = int(raw) - 1
                if 0 <= sel < len(link_targets):
                    open_note_target(data, link_targets[sel]["note"])
        elif ch == "0": return
    while True:
        clear()
        draw_header(f"📄 Note #{note.get('id', 0)}")
        display_note_full(note)
        draw_inline_menu([
            ("1", "Edit"), ("2", "Append"), ("3", "Pin/Unpin"),
            ("4", "Duplicate"), ("5", "Archive"), ("6", "Delete"), ("0", "Back"),
        ])
        ch = draw_prompt()
        if ch == "1":   edit_note(data, note)
        elif ch == "2": append_to_note(data, note)
        elif ch == "3": toggle_pin(data, note)
        elif ch == "4": duplicate_note(data, note); return
        elif ch == "5": archive_note(data, note); return
        elif ch == "6":
            if trash_note(data, note):
                return
        elif ch == "0": return


def edit_note(data: Dict[str, Any], note: Dict[str, Any]) -> None:
    clear()
    draw_header(f"✏️  Edit: {note['title']}")
    pushed_undo = False
    new_title = input(f"    Title [{note['title']}]: ").strip()
    if new_title:
        push_undo(data, f"Edit note #{note['id']}")
        pushed_undo = True
        note["title"] = new_title
    cats = data.get("categories", [])
    if cats:
        print(f"    Categories: {'  '.join(c(f'[{ct}]', cat_color(ct)) for ct in cats)}")
    new_cat = input(f"    Category [{note.get('category', 'General')}]: ").strip()
    if new_cat:
        if not pushed_undo:
            push_undo(data, f"Edit note #{note['id']}")
            pushed_undo = True
        note["category"] = new_cat
        if new_cat not in cats:
            cats.append(new_cat)
    current_tags = format_tags(note.get("tags", []))
    raw_tags = input(f"    Tags [{current_tags or 'none'}]: ").strip()
    if raw_tags:
        if not pushed_undo:
            push_undo(data, f"Edit note #{note['id']}")
            pushed_undo = True
        note["tags"] = parse_tags(raw_tags)
    print(f"\n    Edit body? (y/n) [n]: ", end="")
    print(f"\n    Edit body? (y/n) [n]: ", end="")
    if input().strip().lower() == "y":
        print(c("\n    Current body:", "90"))
        print(wrap_text(note.get("body", ""), indent="      "))
        draw_inline_menu([("1", "Rewrite from scratch"), ("2", "Edit (load existing)"), ("0", "Keep")])
        ec = draw_prompt()
        if ec == "1":
            if not pushed_undo:
                push_undo(data, f"Rewrite #{note['id']}")
                pushed_undo = True
            body = body_input(existing="", hint=True, settings=data.get("settings", {}))
            body = multiline_input()
            if body != "__CANCEL__":
                note["body"] = body
        elif ec == "2":
            if not pushed_undo:
                push_undo(data, f"Edit body #{note['id']}")
                pushed_undo = True
            body = body_input(existing=note.get("body", ""), hint=True, settings=data.get("settings", {}))
            body = multiline_input(existing=note.get("body", ""))
            if body != "__CANCEL__":
                note["body"] = body
    note["updated_at"] = now_iso()
    save_data(data)
    print(f"\n    {c('✓ Updated', '1;32')}")
    pause()


def append_to_note(data: Dict[str, Any], note: Dict[str, Any]) -> None:
    clear()
    draw_header(f"📎 Append to: {note['title']}")
    print(c("    Add text to the end of this note:\n", "90"))
    body = body_input(existing="", hint=True, settings=data.get("settings", {}))
    if body == "__CANCEL__" or not body.strip():
        print("    Cancelled.")
        pause()
        return
    push_undo(data, f"Append to #{note['id']}")
    sep = f"\n\n--- {now_iso()[:16]} ---\n\n"
    existing = note.get("body", "")
    note["body"] = (existing + sep + body) if existing else body
    note["updated_at"] = now_iso()
    save_data(data)
    print(f"\n    {c('✓ Appended', '1;32')}")
    pause()


def toggle_pin(data: Dict[str, Any], note: Dict[str, Any]) -> None:
    note["pinned"] = not note.get("pinned", False)
    save_data(data)
    status = "pinned 📌" if note["pinned"] else "unpinned"
    print(f"    {c(f'✓ Note {status}', '1;32')}")
    time_module.sleep(0.6)


def duplicate_note(data: Dict[str, Any], note: Dict[str, Any]) -> None:
    new_note = {
        "id": next_id(data), "title": note["title"] + " (copy)",
        "body": note.get("body", ""), "category": note.get("category", "General"),
        "tags": list(note.get("tags", [])),
        "created_at": now_iso(), "updated_at": now_iso(), "pinned": False,
    }
    push_undo(data, f"Duplicate #{note['id']}")
    data["notes"].append(new_note)
    save_data(data)
    print(f"    {c('✓ Duplicated', '1;32')} — new note #{new_note['id']}")
    pause()


def archive_note(data: Dict[str, Any], note: Dict[str, Any]) -> None:
    confirm = input(f"    Archive \"{note['title']}\"? (y/n): ").strip().lower()
    if confirm != "y":
        return
    push_undo(data, f"Archive #{note['id']}")
    note["archived_at"] = now_iso()
    data.setdefault("archive", []).append(note)
    data["notes"] = [n for n in data["notes"] if n.get("id") != note["id"]]
    save_data(data)
    print(f"    {c('✓ Archived', '1;32')}")
    pause()


def trash_note(data: Dict[str, Any], note: Dict[str, Any]) -> bool:
    days = data.get("settings", {}).get("trash_days", 30)
    print(c(f"    This moves the note to Trash (recoverable for {days} days).", "33"))
    confirm = input("    Proceed? (y/n): ").strip().lower()
    if confirm != "y":
        return False
    push_undo(data, f"Trash #{note['id']}")
    note["trashed_at"] = now_iso()
    data.setdefault("trash", []).append(note)
    data["notes"] = [n for n in data["notes"] if n.get("id") != note["id"]]
    save_data(data)
    print(f"    {c('✓ Moved to Trash', '1;32')}")
    pause()
    return True


# ════════════════════════════════════════════════════════════════
#  BROWSING & SEARCH
# ════════════════════════════════════════════════════════════════

SORT_MODES = [
    ("recent",   "Most recent first"),
    ("oldest",   "Oldest first"),
    ("alpha",    "A → Z by title"),
    ("alpha_r",  "Z → A by title"),
    ("words",    "Most words first"),
    ("category", "By category"),
]

def sort_notes(notes: List[Dict], mode: str = "recent", pinned_first: bool = True) -> List[Dict]:
    def key(n):
        pin = 1 if (pinned_first and n.get("pinned")) else 0
        ts = n.get("updated_at") or n.get("created_at") or ""
        if mode == "recent":
            return (pin, ts)
        elif mode == "oldest":
            return (pin, "")  # we reverse at end
        elif mode == "alpha":
            return (pin, (n.get("title") or "").lower())
        elif mode == "alpha_r":
            return (pin, (n.get("title") or "").lower())
        elif mode == "words":
            wc = len((n.get("body") or "").split())
            return (pin, wc)
        elif mode == "category":
            return (pin, (n.get("category") or "").lower(), ts)
        return (pin, ts)

    reverse = mode in ("recent", "alpha_r", "words")
    result = sorted(notes, key=key, reverse=reverse)

    if mode == "oldest":
        # Pinned first, then oldest
        pinned = [n for n in notes if n.get("pinned")] if pinned_first else []
        unpinned = [n for n in notes if not n.get("pinned")] if pinned_first else notes
        pinned_sorted = sorted(pinned, key=lambda n: n.get("created_at") or "")
        unpinned_sorted = sorted(unpinned, key=lambda n: n.get("created_at") or "")
        result = pinned_sorted + unpinned_sorted

    return result


def open_note_by_id(data: Dict[str, Any], raw: str) -> None:
    if not raw.isdigit():
        return
    nid = int(raw)
    note = next((n for n in data["notes"] if n.get("id") == nid), None)
    if note:
        view_note(data, note)
        return
    note = next((n for n in data.get("archive", []) if n.get("id") == nid), None)
    if note:
        view_archived_note(data, note)
        return
    note = next((n for n in data.get("trash", []) if n.get("id") == nid), None)
    if note:
        view_trashed_note(data, note)
        return
    print("    Not found.")
    pause()


def browse_notes(data: Dict[str, Any]) -> None:
def browse_notes(data: Dict[str, Any]) -> None:
    cat_filter: Optional[str] = None
    sort_mode = "recent"
    page = 0
    page_size = 10

    while True:
        clear()
        notes = data.get("notes", [])
        filtered = notes[:]
        if cat_filter:
            filtered = [n for n in filtered if (n.get("category") or "").lower() == cat_filter.lower()]
        sorted_n = sort_notes(filtered, mode=sort_mode)
        total = len(sorted_n)
        total_pages = max(1, (total + page_size - 1) // page_size)
        page = min(page, total_pages - 1)
        page_notes = sorted_n[page * page_size:(page + 1) * page_size]

        title = f"📓 All Notes ({total})"
        if cat_filter:
            title += f" — [{cat_filter}]"
        sort_label = next((l for k, l in SORT_MODES if k == sort_mode), "")
        draw_header(title, f"Page {page+1} of {total_pages}  ·  Sort: {sort_label}")

        if not page_notes:
            print(c("    No notes yet. Create one!\n", "90"))
        else:
            for note in page_notes:
                print(format_note_line(note, show_preview=True))
                print()

        opts = [
            ("1", "New note"),      ("5", "Clear filter"),
            ("2", "Open note by #"),("6", "Sort"),
            ("3", "Search"),        ("7", "─"),
            ("4", "Filter category"),
        ]
        if total_pages > 1:
            opts[5] = ("8", "Next page →") if page < total_pages - 1 else ("8", c("─", "90"))
            opts[6] = ("9", "← Prev page") if page > 0 else ("9", c("─", "90"))
        opts.append(("0", "Back"))
        # Clean up empty slots
        opts = [(n, l) for n, l in opts if l != "─" and l != c("─", "90")]
        draw_menu(opts, columns=2)

        ch = draw_prompt()
        if ch == "1":   create_note(data)
        elif ch == "2":
            raw = input("    Note #: ").strip()
            open_note_by_id(data, raw)
        elif ch == "3": search_notes(data)
        elif ch == "4":
            cats = data.get("categories", [])
            if cats:
                print(f"    {'  '.join(c(f'[{ct}]', cat_color(ct)) for ct in cats)}")
            cat_filter = input("    Category (blank to clear): ").strip() or None
            page = 0
        elif ch == "5":
            cat_filter = None
            page = 0
        elif ch == "6":
            # Sort picker
            print()
            for i, (key, label) in enumerate(SORT_MODES, 1):
                marker = c(" ●", "36") if key == sort_mode else ""
                print(f"    {c(f'[{i}]', '36')}  {label}{marker}")
            print()
            sc = draw_prompt()
            if sc.isdigit() and 1 <= int(sc) <= len(SORT_MODES):
                sort_mode = SORT_MODES[int(sc) - 1][0]
                page = 0
        elif ch == "8" and page < total_pages - 1: page += 1
        elif ch == "9" and page > 0: page -= 1
        elif ch == "0": return


def highlight_matches(text: str, keywords: List[str]) -> str:
    highlighted = text
    for kw in keywords:
        highlighted = re.sub(re.escape(kw), lambda m: c(m.group(), "1;33"), highlighted, flags=re.IGNORECASE)
    return highlighted


def search_notes(data: Dict[str, Any]) -> None:
    clear()
    draw_header("🔍 Search Notes", "Searches titles & bodies of all notes including archived")
    query = input("    Search: ").strip()
    if not query:
        return
    keywords = query.lower().split()
    all_notes = data.get("notes", []) + data.get("archive", [])
    results = []
    for note in all_notes:
        tags = " ".join(note.get("tags", []))
        searchable = f"{(note.get('title') or '')} {(note.get('body') or '')} {(note.get('category') or '')} {tags}".lower()
        if all(kw in searchable for kw in keywords):
            score = sum(searchable.count(kw) for kw in keywords)
            if all(kw in (note.get("title") or "").lower() for kw in keywords):
                score += 10
            results.append((note, score, bool(note.get("archived_at"))))
    results.sort(key=lambda x: -x[1])

    clear()
    mw = "match" if len(results) == 1 else "matches"
    draw_header(f"🔍 Results for \"{query}\"", f"{len(results)} {mw}")
    if not results:
        print(c("    No matches found.\n", "90"))
        pause()
        return
    for note, score, is_arch in results[:20]:
        arch_tag = c(" [ARCHIVED]", "90") if is_arch else ""
        print(format_note_line(note, show_preview=False) + arch_tag)
        title = note.get("title", "")
        category = note.get("category", "")
        tags = format_tags(note.get("tags", []))
        title_match = any(kw in title.lower() for kw in keywords)
        cat_match = any(kw in category.lower() for kw in keywords)
        tag_match = any(kw in tags.lower() for kw in keywords)
        if title_match or cat_match or tag_match:
            parts = []
            if title_match:
                parts.append(f"Title: {highlight_matches(title, keywords)}")
            if cat_match:
                parts.append(f"Category: {highlight_matches(category, keywords)}")
            if tag_match:
                parts.append(f"Tags: {highlight_matches(tags, keywords)}")
            print(f"           {c(' · ', '90').join(parts)}")
        body = note.get("body", "")
        if body:
            for kw in keywords:
                idx = body.lower().find(kw)
                if idx >= 0:
                    start = max(0, idx - 30)
                    end = min(len(body), idx + len(kw) + 30)
                    snippet = body[start:end].replace("\n", " ")
                    if start > 0: snippet = "…" + snippet
                    if end < len(body): snippet += "…"
                    highlighted = highlight_matches(snippet, [kw])
                    print(f"           {highlighted}")
                    break
        print()
    draw_inline_menu([("1", "Open note by #"), ("0", "Back")])
    ch = draw_prompt()
    if ch == "1":
        raw = input("    Note #: ").strip()
        open_note_by_id(data, raw)


def restore_archived_note(data: Dict[str, Any], note: Dict[str, Any]) -> None:
    push_undo(data, f"Restore #{note['id']}")
    data["archive"] = [n for n in data["archive"] if n.get("id") != note["id"]]
    note.pop("archived_at", None)
    data["notes"].append(note)
    save_data(data)


def restore_trashed_note(data: Dict[str, Any], note: Dict[str, Any]) -> None:
    push_undo(data, f"Restore #{note['id']} from trash")
    data["trash"] = [n for n in data["trash"] if n.get("id") != note["id"]]
    note.pop("trashed_at", None)
    data["notes"].append(note)
    save_data(data)


def view_archived_note(data: Dict[str, Any], note: Dict[str, Any]) -> None:
    clear()
    draw_header("📦 Archived Note")
    display_note_full(note)
    draw_inline_menu([("1", "Restore to active"), ("2", "Edit (restore first)"), ("0", "Back")])
    ch = draw_prompt()
    if ch == "1":
        restore_archived_note(data, note)
        print(f"    {c('✓ Restored', '1;32')}")
        pause()
    elif ch == "2":
        confirm = input("    Restore and edit this note? (y/n): ").strip().lower()
        if confirm == "y":
            restore_archived_note(data, note)
            view_note(data, note)
    draw_inline_menu([("1", "Restore to active"), ("0", "Back")])
    if draw_prompt() == "1":
        push_undo(data, f"Restore #{note['id']}")
        data["archive"] = [n for n in data["archive"] if n.get("id") != note["id"]]
        note.pop("archived_at", None)
        data["notes"].append(note)
        save_data(data)
        print(f"    {c('✓ Restored', '1;32')}")
        pause()


def view_trashed_note(data: Dict[str, Any], note: Dict[str, Any]) -> None:
    clear()
    draw_header("🗑️  Trashed Note")
    display_note_full(note)
    draw_inline_menu([("1", "Restore to active"), ("2", "Edit (restore first)"), ("0", "Back")])
    ch = draw_prompt()
    if ch == "1":
        restore_trashed_note(data, note)
        print(f"    {c('✓ Restored', '1;32')}")
        pause()
    elif ch == "2":
        confirm = input("    Restore and edit this note? (y/n): ").strip().lower()
        if confirm == "y":
            restore_trashed_note(data, note)
            view_note(data, note)
    draw_inline_menu([("1", "Restore to active"), ("0", "Back")])
    if draw_prompt() == "1":
        push_undo(data, f"Restore #{note['id']} from trash")
        data["trash"] = [n for n in data["trash"] if n.get("id") != note["id"]]
        note.pop("trashed_at", None)
        data["notes"].append(note)
        save_data(data)
        print(f"    {c('✓ Restored', '1;32')}")
        pause()


# ════════════════════════════════════════════════════════════════
#  BROWSE BY DATE
# ════════════════════════════════════════════════════════════════

def browse_by_date(data: Dict[str, Any]) -> None:
    all_notes = data.get("notes", []) + data.get("archive", [])
    if not all_notes:
        clear()
        draw_header("📅 Browse by Date")
        print(c("    No notes yet.\n", "90"))
        pause()
        return
    by_date: Dict[str, List[Dict]] = {}
    for note in all_notes:
        d = (note.get("created_at") or "")[:10]
        if d:
            by_date.setdefault(d, []).append(note)
    sorted_dates = sorted(by_date.keys(), reverse=True)
    page = 0
    dpp = 7
    while True:
        clear()
        tp = max(1, (len(sorted_dates) + dpp - 1) // dpp)
        page = min(page, tp - 1)
        page_dates = sorted_dates[page * dpp:(page + 1) * dpp]
        draw_header("📅 Browse by Date", f"Page {page+1} of {tp}  ·  {len(sorted_dates)} dates")
        for d in page_dates:
            day_notes = by_date[d]
            try:
                dt = date.fromisoformat(d)
                nice = dt.strftime("%B %d, %Y")
                if d == today_str():
                    dl = c(f"📌 Today — {nice}", "1;33")
                elif d == (date.today() - timedelta(days=1)).isoformat():
                    dl = c(f"   Yesterday — {nice}", "37")
                else:
                    dl = c(f"   {dt.strftime('%A')} — {nice}", "37")
            except ValueError:
                dl = c(f"   {d}", "37")
            ac = sum(1 for n in day_notes if n.get("archived_at"))
            cs = f"{len(day_notes)} note{'s' if len(day_notes) != 1 else ''}"
            if ac: cs += f" ({ac} archived)"
            print(f"    {dl}  {c(cs, '90')}")
            for note in sorted(day_notes, key=lambda n: n.get("created_at", ""), reverse=True)[:5]:
                arch = c(" [archived]", "90") if note.get("archived_at") else ""
                pin = c("📌", "33") if note.get("pinned") else "  "
                nc = note.get("category", "")
                nid = note.get("id", 0)
                print(f"      {pin} {c(f'#{nid}', '1;37')}  {c(f'[{nc}]', cat_color(nc))}  {note.get('title', 'Untitled')}{arch}")
            if len(day_notes) > 5:
                print(c(f"      … and {len(day_notes) - 5} more", "90"))
            print()
        opts = [("1", "Open note by #")]
        if page < tp - 1: opts.append(("8", "Next page →"))
        if page > 0: opts.append(("9", "← Prev page"))
        opts.append(("0", "Back"))
        draw_inline_menu(opts)
        ch = draw_prompt()
        if ch == "1":
            open_note_by_id(data, input("    Note #: ").strip())
        elif ch == "8" and page < tp - 1: page += 1
        elif ch == "9" and page > 0: page -= 1
        elif ch == "0": return


# ════════════════════════════════════════════════════════════════
#  ARCHIVE & TRASH BROWSERS
# ════════════════════════════════════════════════════════════════

def archive_browser(data: Dict[str, Any]) -> None:
    page = 0
    ps = 12
    while True:
        clear()
        arch = data.get("archive", [])
        sa = sorted(arch, key=lambda n: n.get("archived_at", ""), reverse=True)
        total = len(sa)
        tp = max(1, (total + ps - 1) // ps)
        page = min(page, tp - 1)
        pn = sa[page * ps:(page + 1) * ps]
        draw_header(f"📦 Archive ({total} notes)", f"Page {page+1} of {tp}")
        if not pn:
            print(c("    Archive is empty.\n", "90"))
            pause()
            return
        for note in pn:
            ad = (note.get("archived_at") or "")[:10]
            nc = note.get("category", "")
            nid = note.get("id", 0)
            print(f"    {c(f'#{nid:<4}', '1;37')}  {c(ad, '90')}  {c(f'[{nc}]', cat_color(nc))}  {note.get('title', 'Untitled')}")
        opts = [("1", "View note"), ("2", "Restore note")]
        if page < tp - 1: opts.append(("8", "Next page →"))
        if page > 0: opts.append(("9", "← Prev page"))
        opts.append(("0", "Back"))
        draw_inline_menu(opts)
        ch = draw_prompt()
        if ch == "1":
            raw = input("    Note #: ").strip()
            if raw.isdigit():
                n = next((x for x in arch if x.get("id") == int(raw)), None)
                if n: view_archived_note(data, n)
                else: print("    Not found."); pause()
        elif ch == "2":
            raw = input("    Note # to restore: ").strip()
            if raw.isdigit():
                nid = int(raw)
                for i, n in enumerate(arch):
                    if n.get("id") == nid:
                        restore_archived_note(data, n)
                        print(f"    {c('✓ Restored', '1;32')}")
                        pause()
                        break
                else:
                    print("    Not found."); pause()
        elif ch == "8" and page < tp - 1: page += 1
        elif ch == "9" and page > 0: page -= 1
        elif ch == "0": return


def trash_browser(data: Dict[str, Any]) -> None:
    page = 0
    ps = 12
    while True:
        clear()
        trash = data.get("trash", [])
        st = sorted(trash, key=lambda n: n.get("trashed_at", ""), reverse=True)
        total = len(st)
        days = data.get("settings", {}).get("trash_days", 30)
        tp = max(1, (total + ps - 1) // ps)
        page = min(page, tp - 1)
        pn = st[page * ps:(page + 1) * ps]
        draw_header(f"🗑️  Trash ({total} notes)", f"Auto-purged after {days} days  ·  Page {page+1} of {tp}")
        if not pn:
            print(c("    Trash is empty.\n", "90"))
            pause()
            return
        for note in pn:
            td = (note.get("trashed_at") or "")[:10]
            nid = note.get("id", 0)
            # Days remaining
            try:
                trashed_dt = datetime.fromisoformat(note["trashed_at"])
                expires = trashed_dt + timedelta(days=days)
                remaining = (expires - datetime.now()).days
                exp_str = c(f"{remaining}d left", "33") if remaining > 7 else c(f"{remaining}d left", "31")
            except Exception:
                exp_str = ""
            print(f"    {c(f'#{nid:<4}', '1;37')}  {c(td, '90')}  {exp_str}  {note.get('title', 'Untitled')}")
        opts = [("1", "Restore note"), ("2", "Empty trash"), ("3", "View note")]
        if page < tp - 1: opts.append(("8", "Next page →"))
        if page > 0: opts.append(("9", "← Prev page"))
        opts.append(("0", "Back"))
        draw_inline_menu(opts)
        ch = draw_prompt()
        if ch == "1":
            raw = input("    Note # to restore: ").strip()
            if raw.isdigit():
                nid = int(raw)
                for i, n in enumerate(trash):
                    if n.get("id") == nid:
                        restore_trashed_note(data, n)
                        print(f"    {c('✓ Restored', '1;32')}")
                        pause()
                        break
                else:
                    print("    Not found."); pause()
        elif ch == "2":
            confirm = input(f"    Permanently delete all {total} notes? Type EMPTY: ").strip()
            if confirm == "EMPTY":
                push_undo(data, f"Empty trash ({total} notes)")
                data["trash"] = []
                save_data(data)
                print(f"    {c('✓ Trash emptied', '1;32')}")
                pause()
                return
        elif ch == "3":
            raw = input("    Note #: ").strip()
            if raw.isdigit():
                nid = int(raw)
                n = next((x for x in trash if x.get("id") == nid), None)
                if n:
                    view_trashed_note(data, n)
                else:
                    print("    Not found."); pause()
        elif ch == "8" and page < tp - 1: page += 1
        elif ch == "9" and page > 0: page -= 1
        elif ch == "0": return


# ════════════════════════════════════════════════════════════════
#  QUICK NOTE
# ════════════════════════════════════════════════════════════════

def quick_note(data: Dict[str, Any]) -> None:
    clear()
    draw_header("⚡ Quick Note", "Fast capture — just type and go")
    text = input(f"    {c('›', '33')} ").strip()
    if not text:
        return
    if len(text) <= 80:
        title, body = text, ""
    else:
        parts = text.split(". ", 1)
        title = parts[0]
        body = parts[1] if len(parts) > 1 else ""
    note = {
        "id": next_id(data), "title": title, "body": body,
        "category": data.get("settings", {}).get("default_category", "General"),
        "tags": [],
        "created_at": now_iso(), "updated_at": now_iso(), "pinned": False,
    }
    push_undo(data, f"Quick note #{note['id']}")
    data.setdefault("notes", []).append(note)
    save_data(data)
    print(f"\n    {c('✓ Captured', '1;32')} — #{note['id']} \"{title}\"")
    pause()


# ════════════════════════════════════════════════════════════════
#  STATS
# ════════════════════════════════════════════════════════════════

def show_stats(data: Dict[str, Any]) -> None:
    clear()
    notes = data.get("notes", [])
    archive = data.get("archive", [])
    trash = data.get("trash", [])
    all_notes = notes + archive
    tw = sum(len((n.get("body") or "").split()) for n in all_notes)
    tc = sum(len(n.get("body") or "") for n in all_notes)

    draw_header("📊 Notes Stats")
    draw_section("Overview")
    print(f"    Active notes     {c(str(len(notes)), '1;37')}")
    print(f"    Archived         {c(str(len(archive)), '90')}")
    print(f"    In trash         {c(str(len(trash)), '90')}")
    print(f"    Total written    {c(str(len(all_notes)), '1;37')}")
    print(f"    Total words      {c(f'{tw:,}', '33')}")
    print(f"    Total chars      {c(f'{tc:,}', '90')}")
    print()

    cats: Dict[str, int] = {}
    for n in all_notes:
        cat = n.get("category") or "Uncategorized"
        cats[cat] = cats.get(cat, 0) + 1
    if cats:
        draw_section("By Category")
        mx = max(cats.values())
        for cat, cnt in sorted(cats.items(), key=lambda x: -x[1]):
            bl = round((cnt / mx) * 25) if mx else 0
            bar = c("█" * bl, cat_color(cat)) + c("░" * (25 - bl), "90")
            print(f"    {cat:<15} {bar}  {cnt}")
        print()

    tags: Dict[str, int] = {}
    for n in all_notes:
        for tag in n.get("tags", []):
            tags[tag] = tags.get(tag, 0) + 1
    if tags:
        draw_section("By Tag")
        mx = max(tags.values())
        for tag, cnt in sorted(tags.items(), key=lambda x: -x[1])[:15]:
            bl = round((cnt / mx) * 25) if mx else 0
            bar = c("█" * bl, "36") + c("░" * (25 - bl), "90")
            print(f"    {tag:<15} {bar}  {cnt}")
        print()

    months: Dict[str, int] = {}
    for n in all_notes:
        m = (n.get("created_at") or "")[:7]
        if m: months[m] = months.get(m, 0) + 1
    if months:
        draw_section("By Month")
        mx = max(months.values())
        for month in sorted(months.keys(), reverse=True)[:12]:
            cnt = months[month]
            bl = round((cnt / mx) * 25) if mx else 0
            bar = c("█" * bl, "36") + c("░" * (25 - bl), "90")
            print(f"    {month}  {bar}  {cnt}")
        print()

    pinned = sum(1 for n in notes if n.get("pinned"))
    if pinned:
        print(f"    📌 Pinned notes: {pinned}\n")

    # Storage info
    draw_section("Storage")
    dp = data_path()
    size_str = "N/A"
    if os.path.exists(dp):
        size = os.path.getsize(dp)
        if size < 1024:
            size_str = f"{size} B"
        elif size < 1024 * 1024:
            size_str = f"{size / 1024:.1f} KB"
        else:
            size_str = f"{size / (1024*1024):.1f} MB"
    print(f"    Data file   {c(dp, '90')}")
    print(f"    File size   {c(size_str, '90')}")
    print(f"    Export dir  {c(export_dir(), '90')}")
    print()
    pause()


# ════════════════════════════════════════════════════════════════
#  EXPORT
# ════════════════════════════════════════════════════════════════

def export_menu(data: Dict[str, Any]) -> None:
    clear()
    draw_header("📤 Export Notes")
    draw_menu([
        ("1", "Export all to Markdown"),
        ("2", "Export single note to Markdown"),
        ("3", "Export all to plain text"),
        ("4", "Export all to CSV"),
        ("5", "Export all to Excel (.xlsx)"),
        ("6", "Export filtered (status/category/date)"),
        ("0", "Back"),
    ], columns=1)
    ch = draw_prompt()
    edir = export_dir()
    os.makedirs(edir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if ch == "1":
        path = os.path.join(edir, f"notes_all_{ts}.md")
        export_all_markdown(data, path)
        print(f"\n    {c('✓ Saved:', '1;32')} {path}")
        pause()
    elif ch == "2":
        raw = input("    Note #: ").strip()
        if raw.isdigit():
            nid = int(raw)
            all_n = data.get("notes", []) + data.get("archive", [])
            note = next((n for n in all_n if n.get("id") == nid), None)
            if note:
                safe = re.sub(r'[^\w\s-]', '', note.get("title", "note")).strip().replace(" ", "_")[:40]
                path = os.path.join(edir, f"{safe}_{ts}.md")
                export_single_markdown(note, path)
                print(f"\n    {c('✓ Saved:', '1;32')} {path}")
            else:
                print("    Not found.")
        pause()
    elif ch == "3":
        path = os.path.join(edir, f"notes_all_{ts}.txt")
        export_all_text(data, path)
        print(f"\n    {c('✓ Saved:', '1;32')} {path}")
        pause()
    elif ch == "4":
        path = os.path.join(edir, f"notes_{ts}.csv")
        export_csv(data, path)
        print(f"\n    {c('✓ Saved:', '1;32')} {path}")
        pause()
    elif ch == "5":
        path = os.path.join(edir, f"notes_{ts}.xlsx")
        export_excel(data, path)
        print(f"\n    {c('✓ Saved:', '1;32')} {path}")
        pause()
    elif ch == "6":
        export_filtered_menu(data, edir, ts)


def _all_exportable(data: Dict[str, Any]) -> List[Dict]:
    return data.get("notes", []) + data.get("archive", [])


def parse_date_input(raw: str) -> Optional[date]:
    raw = raw.strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return None


def filter_notes(data: Dict[str, Any], status: str, category: Optional[str],
                 start: Optional[date], end: Optional[date]) -> List[Dict[str, Any]]:
    if status == "active":
        notes = data.get("notes", [])
    elif status == "archived":
        notes = data.get("archive", [])
    else:
        notes = _all_exportable(data)
    if category:
        notes = [n for n in notes if (n.get("category") or "").lower() == category.lower()]
    if start or end:
        filtered = []
        for n in notes:
            created = (n.get("created_at") or "")[:10]
            if not created:
                continue
            try:
                created_date = date.fromisoformat(created)
            except ValueError:
                continue
            if start and created_date < start:
                continue
            if end and created_date > end:
                continue
            filtered.append(n)
        notes = filtered
    return notes


def export_filtered_menu(data: Dict[str, Any], edir: str, ts: str) -> None:
    print()
    draw_section("Filter")
    draw_inline_menu([("1", "All"), ("2", "Active only"), ("3", "Archived only")])
    status_choice = draw_prompt()
    status = "all"
    if status_choice == "2":
        status = "active"
    elif status_choice == "3":
        status = "archived"
    cats = data.get("categories", [])
    if cats:
        print(f"    Categories: {'  '.join(c(f'[{ct}]', cat_color(ct)) for ct in cats)}")
    category = input("    Category (blank for all): ").strip() or None
    start_raw = input("    Start date (YYYY-MM-DD, optional): ").strip()
    end_raw = input("    End date (YYYY-MM-DD, optional): ").strip()
    start = parse_date_input(start_raw)
    end = parse_date_input(end_raw)
    notes = filter_notes(data, status, category, start, end)
    if not notes:
        print(c("    No notes match this filter.", "33"))
        pause()
        return
    print()
    draw_inline_menu([("1", "Markdown"), ("2", "Plain text"), ("3", "CSV")])
    fmt = draw_prompt()
    if fmt == "1":
        path = os.path.join(edir, f"notes_filtered_{ts}.md")
        export_markdown_notes(notes, path)
        print(f"\n    {c('✓ Saved:', '1;32')} {path}")
    elif fmt == "2":
        path = os.path.join(edir, f"notes_filtered_{ts}.txt")
        export_text_notes(notes, path)
        print(f"\n    {c('✓ Saved:', '1;32')} {path}")
    elif fmt == "3":
        path = os.path.join(edir, f"notes_filtered_{ts}.csv")
        export_csv_notes(notes, path)
        print(f"\n    {c('✓ Saved:', '1;32')} {path}")
    pause()

def export_markdown_notes(notes: List[Dict[str, Any]], path: str) -> None:
    all_n = notes
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"# Notes Vault Export — {today_str()}\n\nTotal: {len(all_n)} notes\n\n---\n\n")
        cats: Dict[str, List[Dict]] = {}
        for n in sorted(all_n, key=lambda x: x.get("created_at", ""), reverse=True):
            cats.setdefault(n.get("category") or "Uncategorized", []).append(n)
        for cat in sorted(cats):
            f.write(f"## {cat}\n\n")
            for n in cats[cat]:
                pin = "📌 " if n.get("pinned") else ""
                arch = " *(archived)*" if n.get("archived_at") else ""
                f.write(f"### {pin}{n.get('title', 'Untitled')}{arch}\n\n")
                f.write(f"*Created: {n.get('created_at', '')[:16]}")
                if n.get("updated_at") and n["updated_at"] != n.get("created_at"):
                    f.write(f" · Updated: {n['updated_at'][:16]}")
                tags = format_tags(n.get("tags", []))
                if tags:
                    f.write(f" · Tags: {tags}")
                f.write(f"*\n\n{n.get('body', '')}\n\n---\n\n")


def export_all_markdown(data: Dict[str, Any], path: str) -> None:
    export_markdown_notes(_all_exportable(data), path)


def export_single_markdown(note: Dict[str, Any], path: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        pin = "📌 " if note.get("pinned") else ""
        f.write(f"# {pin}{note.get('title', 'Untitled')}\n\n")
        tags = format_tags(note.get("tags", []))
        tags_part = f" · Tags: {tags}" if tags else ""
        f.write(f"*Category: {note.get('category', '')}{tags_part} · Created: {note.get('created_at', '')[:16]}*\n\n")
        f.write(note.get("body", "") + "\n")


def export_text_notes(notes: List[Dict[str, Any]], path: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        for n in sorted(notes, key=lambda x: x.get("created_at", ""), reverse=True):
            f.write(f"{'=' * 60}\n#{n.get('id')} — {n.get('title', 'Untitled')}\n")
            tags = format_tags(n.get("tags", []))
            f.write(f"Category: {n.get('category', '')}  |  Tags: {tags}  |  {n.get('created_at', '')[:16]}\n")
            f.write(f"{'=' * 60}\n\n{n.get('body', '')}\n\n\n")


def export_all_text(data: Dict[str, Any], path: str) -> None:
    export_text_notes(_all_exportable(data), path)


def export_csv_notes(notes: List[Dict[str, Any]], path: str) -> None:
    all_n = notes
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["ID", "Title", "Category", "Tags", "Status", "Pinned", "Created", "Updated",
                    "Word Count", "Char Count", "Body Preview (100 chars)", "Full Body"])
        for n in sorted(all_n, key=lambda x: x.get("created_at", ""), reverse=True):
            body = n.get("body", "")
            wc = len(body.split()) if body else 0
            status = "Archived" if n.get("archived_at") else "Active"
            preview = body.replace("\n", " ")[:100]
            w.writerow([
                n.get("id"), n.get("title", ""), n.get("category", ""), format_tags(n.get("tags", [])),
                status, "Yes" if n.get("pinned") else "No",
                n.get("created_at", "")[:16], (n.get("updated_at") or "")[:16],
                wc, len(body), preview, body,
            ])


def export_csv(data: Dict[str, Any], path: str) -> None:
    export_csv_notes(_all_exportable(data), path)


def export_excel(data: Dict[str, Any], path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print(c("    ⚠ openpyxl not installed. Run: pip install openpyxl", "31"))
        pause()
        return

    all_n = _all_exportable(data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Notes"

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    date_font = Font(name="Arial", size=10, color="666666")
    body_font = Font(name="Arial", size=10)
    pin_font = Font(name="Arial", size=10, color="D4A017", bold=True)
    thin_border = Border(
        bottom=Side(style="thin", color="D9D9D9"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")
    top_align = Alignment(vertical="top")

    # Headers
    headers = ["ID", "Title", "Category", "Tags", "Status", "Pinned", "Created", "Updated",
               "Words", "Chars", "Body"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_idx, n in enumerate(sorted(all_n, key=lambda x: x.get("created_at", ""), reverse=True), 2):
        body = n.get("body", "")
        wc = len(body.split()) if body else 0
        status = "Archived" if n.get("archived_at") else "Active"
        is_pinned = n.get("pinned", False)

        values = [
            n.get("id"), n.get("title", ""), n.get("category", ""), format_tags(n.get("tags", [])),
            status, "📌" if is_pinned else "",
            n.get("created_at", "")[:16], (n.get("updated_at") or "")[:16],
            wc, len(body), body,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = body_font
            cell.alignment = top_align
            cell.border = thin_border

        # Highlight pinned rows
        if is_pinned:
            for col in range(1, len(values) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FFF8E1")

        # Date columns in gray
        for col in (7, 8):
            ws.cell(row=row_idx, column=col).font = date_font

        # Body column wrap
        ws.cell(row=row_idx, column=11).alignment = wrap_align

    # Column widths
    widths = [6, 35, 14, 18, 10, 7, 18, 18, 8, 8, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:K{len(all_n) + 1}"

    wb.save(path)


# ════════════════════════════════════════════════════════════════
#  SETTINGS
# ════════════════════════════════════════════════════════════════

def settings_menu(data: Dict[str, Any]) -> None:
    while True:
        clear()
        settings = data.get("settings", {})
        cats = data.get("categories", [])
        templates = data.get("templates", [])

        draw_header("⚙️  Settings")

        draw_section("Current")
        hint_st = c("ON", "32") if settings.get("editor_hint", True) else c("OFF", "31")
        trash_days = settings.get("trash_days", 30)
        print(f"    Default category   {c(settings.get('default_category', 'General'), '35')}")
        print(f"    Editor hints       {hint_st}")
        ext_editor = settings.get("use_external_editor", False)
        ext_editor_st = c("ON", "32") if ext_editor else c("OFF", "31")
        print(f"    External editor    {ext_editor_st}")
        print(f"    Trash retention    {c(f'{trash_days} days', '33')}")
        print(f"    Categories         {'  '.join(c(f'[{ct}]', cat_color(ct)) for ct in cats)}")
        if templates:
            print(f"    Templates          {', '.join(t['name'] for t in templates)}")
        print()

        draw_section("Storage")
        print(f"    Data file          {c(data_path(), '36')}")
        print(f"    Export directory    {c(export_dir(), '36')}")
        print(f"    Config file        {c(CONFIG_PATH, '90')}")

        draw_menu([
            ("1", "Change default category"),
            ("2", "Toggle editor hints"),
            ("3", "Add category"),
            ("4", "Remove category"),
            ("5", "Change data file location"),
            ("6", "Change export directory"),
            ("7", "Set trash retention days"),
            ("8", "Manage templates"),
            ("9", "Toggle external editor"),
            ("0", "Back"),
        ], columns=1)

        ch = draw_prompt()

        if ch == "1":
            if cats: print(f"    Available: {', '.join(cats)}")
            new = input("    Default category: ").strip()
            if new:
                settings["default_category"] = new
                if new not in cats: cats.append(new)
                save_data(data)

        elif ch == "2":
            settings["editor_hint"] = not settings.get("editor_hint", True)
            save_data(data)

        elif ch == "3":
            name = input("    New category name: ").strip()
            if name and name not in cats:
                cats.append(name)
                data["categories"] = cats
                save_data(data)

        elif ch == "4":
            if cats:
                for i, ct in enumerate(cats, 1):
                    print(f"      {c(f'[{i}]', '36')} {ct}")
            name = input("    Category to remove: ").strip()
            if name in cats:
                affected_notes = [
                    n for n in (data.get("notes", []) + data.get("archive", []) + data.get("trash", []))
                    if (n.get("category") or "").lower() == name.lower()
                ]
                if affected_notes:
                    default_cat = settings.get("default_category", "General")
                    print(c(f"    {len(affected_notes)} notes use \"{name}\".", "33"))
                    replace = input(f"    Replace with category [{default_cat}]: ").strip() or default_cat
                    for n in affected_notes:
                        n["category"] = replace
                    if replace not in cats:
                        cats.append(replace)
                    save_data(data)
                cats.remove(name)
                data["categories"] = cats
                save_data(data)

        elif ch == "5":
            print(f"\n    Current: {c(data_path(), '36')}")
            print(c("    Enter new path (absolute). Data will be moved automatically.", "90"))
            new_path = input("    New path: ").strip()
            if new_path:
                new_path = os.path.expanduser(new_path)
                if not new_path.endswith(".json"):
                    new_path += ".json"
                old_path = data_path()
                try:
                    os.makedirs(os.path.dirname(new_path) or ".", exist_ok=True)
                    if os.path.exists(old_path):
                        shutil.copy2(old_path, new_path)
                        print(f"    {c('✓ Data copied to new location', '32')}")
                    CONFIG["data_path"] = new_path
                    save_config(CONFIG)
                    print(f"    {c('✓ Config updated', '32')}")
                    if os.path.exists(old_path) and old_path != new_path:
                        rm = input(f"    Delete old file at {old_path}? (y/n): ").strip().lower()
                        if rm == "y":
                            os.remove(old_path)
                            print(f"    {c('✓ Old file removed', '32')}")
                except Exception as e:
                    print(f"    {c(f'Error: {e}', '31')}")
                pause()

        elif ch == "6":
            print(f"\n    Current: {c(export_dir(), '36')}")
            new_dir = input("    New export directory: ").strip()
            if new_dir:
                new_dir = os.path.expanduser(new_dir)
                os.makedirs(new_dir, exist_ok=True)
                CONFIG["export_dir"] = new_dir
                save_config(CONFIG)
                print(f"    {c('✓ Export directory updated', '32')}")
                pause()

        elif ch == "7":
            cur = settings.get("trash_days", 30)
            new_days = input(f"    Trash retention days [{cur}]: ").strip()
            if new_days.isdigit() and int(new_days) > 0:
                settings["trash_days"] = int(new_days)
                save_data(data)

        elif ch == "8":
            manage_templates(data)

        elif ch == "9":
            settings["use_external_editor"] = not settings.get("use_external_editor", False)
            save_data(data)

        elif ch == "0":
            return


def manage_templates(data: Dict[str, Any]) -> None:
    while True:
        clear()
        templates = data.get("templates", [])
        draw_header("📝 Note Templates")
        if not templates:
            print(c("    No templates.\n", "90"))
        else:
            for i, t in enumerate(templates, 1):
                preview = t.get("body", "").replace("\n", " ")[:50]
                print(f"    {c(f'[{i}]', '36')}  {c(t['name'], '1;37')}")
                print(f"         {c(preview + '...', '90')}")
            print()
        draw_inline_menu([("1", "Add template"), ("2", "Remove template"), ("3", "Edit template"), ("0", "Back")])
        ch = draw_prompt()
        if ch == "1":
            name = input("    Template name: ").strip()
            if name:
                print(c("\n    Enter template body (this will pre-fill new notes):", "90"))
                body = multiline_input(hint=False)
                if body != "__CANCEL__":
                    templates.append({"name": name, "body": body})
                    data["templates"] = templates
                    save_data(data)
                    print(f"    {c('✓ Template added', '1;32')}")
                    pause()
        elif ch == "2":
            if templates:
                raw = input("    Template number to remove: ").strip()
                if raw.isdigit():
                    idx = int(raw) - 1
                    if 0 <= idx < len(templates):
                        removed = templates.pop(idx)
                        data["templates"] = templates
                        save_data(data)
                        rname = removed.get("name", "")
                        print(f"    {c(f'✓ Removed: {rname}', '1;32')}")
                        pause()
        elif ch == "3":
            if templates:
                raw = input("    Template number to edit: ").strip()
                if raw.isdigit():
                    idx = int(raw) - 1
                    if 0 <= idx < len(templates):
                        tmpl = templates[idx]
                        new_name = input(f"    Name [{tmpl.get('name', '')}]: ").strip()
                        if new_name:
                            tmpl["name"] = new_name
                        print(c("\n    Edit template body (leave blank to keep current):", "90"))
                        body = multiline_input(existing=tmpl.get("body", ""), hint=False)
                        if body != "__CANCEL__":
                            tmpl["body"] = body
                        data["templates"] = templates
                        save_data(data)
                        print(f"    {c('✓ Template updated', '1;32')}")
                        pause()
        elif ch == "0":
            return


# ════════════════════════════════════════════════════════════════
#  MAIN MENU
# ════════════════════════════════════════════════════════════════

def main():
    data = load_data()
    auto_purge_trash(data)

    while True:
        clear()
        notes = data.get("notes", [])
        archive = data.get("archive", [])
        trash = data.get("trash", [])
        pinned = [n for n in notes if n.get("pinned")]
        recent = sort_notes(notes)[:3]

        sub_parts = [f"{len(notes)} notes"]
        if archive: sub_parts.append(f"{len(archive)} archived")
        if trash: sub_parts.append(f"{len(trash)} in trash")
        draw_header("📓  N O T E S   V A U L T", "  ·  ".join(sub_parts))

        if pinned:
            draw_section("📌 Pinned")
            for n in pinned[:4]:
                nid = n.get("id", 0)
                cl = c(f"[{n.get('category', '')}]", cat_color(n.get("category", "")))
                print(f"    {c(f'#{nid}', '1;37')}  {cl}  {n.get('title', 'Untitled')}")
            print()
        elif recent:
            draw_section("Recent")
            for n in recent:
                nid = n.get("id", 0)
                cl = c(f"[{n.get('category', '')}]", cat_color(n.get("category", "")))
                ds = c((n.get("created_at") or "")[:10], "90")
                print(f"    {c(f'#{nid}', '1;37')}  {ds}  {cl}  {n.get('title', 'Untitled')}")
            print()

        draw_menu([
            ("1",  "New note"),         ("7",  "Archive"),
            ("2",  "Quick note"),       ("8",  "Trash"),
            ("3",  "Browse all"),       ("9",  "Export"),
            ("4",  "Search"),           ("10", "Stats"),
            ("5",  "Browse by date"),   ("11", "Settings"),
            ("6",  "Open note by #"),   ("00", "Undo"),
            ("0",  "Exit"),
        ], columns=2)

        ch = draw_prompt()

        if ch == "1":    create_note(data)
        elif ch == "2":  quick_note(data)
        elif ch == "3":  browse_notes(data)
        elif ch == "4":  search_notes(data)
        elif ch == "5":  browse_by_date(data)
        elif ch == "6":
            raw = input("    Note #: ").strip()
            open_note_by_id(data, raw)
        elif ch == "7":  archive_browser(data)
        elif ch == "8":  trash_browser(data)
        elif ch == "9":  export_menu(data)
        elif ch == "10": show_stats(data)
        elif ch == "11": settings_menu(data)
        elif ch == "00":
            desc = do_undo(data)
            if desc:  print(f"    {c('✓ Undone:', '1;32')} {desc}")
            else:     print(c("    Nothing to undo.", "90"))
            pause()
        elif ch == "0":
            save_data(data)
            print(f"\n    {c('See you later! 👋', '90')}\n")
            break


if __name__ == "__main__":
    main()
